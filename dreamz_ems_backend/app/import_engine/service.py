"""Import service (sprint-3/09) — two-phase, server-authoritative pipeline.

Validate (dry-run, zero writes) → Commit (ALL-OR-NOTHING single transaction).
Set-based throughout (one batched query per resolver / match-existence; bulk DML)
— never per-row queries (D7). Owns: create_job (caps fail-fast + quarantine
store), validate, commit, template build, annotated-error build.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

from fastapi import HTTPException, status as http_status
from sqlalchemy.orm import Session

from app.config import settings
from app.import_engine import readers
from app.import_engine.coerce import coerce, to_utc
from app.import_engine.registry import (
    ImportColumn,
    ImporterDef,
    RESOLVE_FIND_OR_CREATE,
    get_importer,
)
from app.import_engine.sanitize import sanitize_cell
from app.models.import_job import (
    IMPORT_MODES,
    ImportJob,
    MODE_CREATE_ONLY,
    MODE_UPDATE_ONLY,
    MODE_UPSERT,
    STATUS_DONE,
    STATUS_FAILED,
    STATUS_PENDING,
    STATUS_VALIDATED,
    STATUS_VALIDATING,
)
from app.repositories.import_repository import ImportRepository
from app.services.storage import storage_for_tenant

logger = logging.getLogger(__name__)

ERRORS_CAP = 1000  # on-page / errors_json cap (full set in the error file)

_EXT = {readers.FMT_XLSX: "xlsx", readers.FMT_XLS: "xls", readers.FMT_CSV: "csv"}


class PreparedRow:
    __slots__ = ("index", "data", "record_id", "op")

    def __init__(self, index: int, data: dict, record_id: Optional[str], op: str):
        self.index = index
        self.data = data
        self.record_id = record_id
        self.op = op  # 'create' | 'update'


class ImportService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = ImportRepository(db)

    # ── caps ────────────────────────────────────────────────────────────────

    def caps(self, tenant_id: str) -> Tuple[int, int]:
        row = self.repo.get_settings(tenant_id)
        max_rows = (row.max_rows if row and row.max_rows else None) or settings.import_max_rows
        max_mb = (row.max_file_mb if row and row.max_file_mb else None) or settings.import_max_file_mb
        return max_rows, max_mb

    def get_settings(self, tenant_id: str) -> dict:
        """Effective caps + whether they're the global default (D11)."""
        row = self.repo.get_settings(tenant_id)
        is_default = not (row and (row.max_rows or row.max_file_mb))
        max_rows, max_mb = self.caps(tenant_id)
        return {"maxRows": max_rows, "maxFileMb": max_mb, "isDefault": is_default}

    def set_settings(
        self, tenant_id: str, max_rows: Optional[int], max_file_mb: Optional[int]
    ) -> dict:
        """Set per-tenant caps. None/0 on a field falls back to the global default."""
        self.repo.upsert_settings(
            tenant_id,
            max_rows if max_rows and max_rows > 0 else None,
            max_file_mb if max_file_mb and max_file_mb > 0 else None,
        )
        self.db.commit()
        return self.get_settings(tenant_id)

    # ── create (upload, fail-fast caps, quarantine store) ────────────────────

    def create_job(
        self,
        *,
        tenant_id: str,
        actor_user_id: str,
        entity_type: str,
        mode: str,
        abort_on_invalid: bool,
        trigger_automations: bool,
        context: Optional[dict],
        content: bytes,
        ext_hint: str,
    ) -> ImportJob:
        importer = self._require_importer(entity_type)
        if mode not in IMPORT_MODES:
            raise HTTPException(422, f"Unknown mode: {mode}")
        max_rows, max_mb = self.caps(tenant_id)
        if len(content) > max_mb * 1024 * 1024:
            raise HTTPException(413, f"File exceeds the {max_mb} MB limit.")
        fmt = readers.sniff_format(content)
        if fmt is None:
            raise HTTPException(422, "Unsupported file — upload xlsx, xls or csv.")

        ctx = {k: v for k, v in (context or {}).items() if k in importer.context_keys}
        job = ImportJob(
            tenant_id=tenant_id,
            actor_user_id=actor_user_id,
            entity_type=entity_type,
            mode=mode,
            abort_on_invalid=abort_on_invalid,
            trigger_automations=trigger_automations,
            context_json=ctx or None,
            status=STATUS_PENDING,
        )
        self.repo.add(job)  # need the id for the quarantine key
        store = storage_for_tenant(self.db, tenant_id)
        ext = _EXT.get(fmt, ext_hint or "dat")
        job.file_storage_key = store.save(
            f"imports/{job.id}/source.{ext}", content, "application/octet-stream"
        )
        self.db.commit()
        return job

    # ── mapping + enqueue validate ───────────────────────────────────────────

    def set_mapping(
        self,
        tenant_id: str,
        job_id: str,
        mapping: dict,
        sheet_name: Optional[str],
        *,
        context: Optional[dict] = None,
    ) -> ImportJob:
        job = self._require_job(tenant_id, job_id)
        job.mapping_json = mapping or {}
        job.sheet_name = sheet_name
        if context:
            self._merge_context(job, context)
        job.status = STATUS_VALIDATING
        self.db.commit()
        self._enqueue("validate", job.id)
        return job

    def _merge_context(self, job: ImportJob, context: dict) -> None:
        """Merge whitelisted job-level context (e.g. Ticket mode) onto the job.
        Keys are constrained to the importer's ``context_keys`` (client can't
        widen the scope) — mirrors create_job's filter."""
        importer = self._require_importer(job.entity_type)
        allowed = {k: v for k, v in (context or {}).items() if k in importer.context_keys}
        if allowed:
            job.context_json = {**(job.context_json or {}), **allowed}

    def commit_job(
        self,
        tenant_id: str,
        job_id: str,
        *,
        abort_on_invalid: Optional[bool] = None,
        trigger_automations: Optional[bool] = None,
        context: Optional[dict] = None,
    ) -> ImportJob:
        job = self._require_job(tenant_id, job_id)
        if job.status != STATUS_VALIDATED:
            raise HTTPException(409, "Job is not ready to commit.")
        # The Abort + Trigger-automations choices live on the import page now,
        # applied here at commit (override whatever the job was created with).
        if abort_on_invalid is not None:
            job.abort_on_invalid = abort_on_invalid
        if trigger_automations is not None:
            job.trigger_automations = trigger_automations
        if context:
            self._merge_context(job, context)
        self.db.commit()
        self._enqueue("commit", job.id)
        return job

    def _enqueue(self, phase: str, job_id: str) -> None:
        """Eager (dev/test) runs inline on THIS session; else Celery .delay."""
        if settings.celery_task_always_eager:
            if phase == "validate":
                self.run_validate(self.db, job_id)
            else:
                self.run_commit(self.db, job_id)
            return
        from app.import_engine.worker import import_commit_task, import_validate_task

        (import_validate_task if phase == "validate" else import_commit_task).delay(job_id)

    # ── shared row preparation (mapping → coerce → validate → resolve) ────────

    def _read_content(self, job: ImportJob) -> bytes:
        store = storage_for_tenant(self.db, job.tenant_id)
        kind, value = store.resolve(job.file_storage_key)
        if kind == "path":
            with open(value, "rb") as fh:
                return fh.read()
        import urllib.request

        return urllib.request.urlopen(value).read()  # noqa: S310 (own storage)

    def _invert_mapping(self, importer: ImporterDef, mapping: dict):
        """columnKey → fileHeader. Map collisions DEGRADE (D4): 2 file-cols → 1
        system-col leaves the target blank + a warning (never blocks)."""
        targets: Dict[str, List[str]] = {}
        for file_header, col_key in (mapping or {}).items():
            if not col_key:  # "Don't import"
                continue
            targets.setdefault(col_key, []).append(file_header)
        resolved: Dict[str, Optional[str]] = {}
        warnings: List[str] = []
        for col_key, headers in targets.items():
            if len(headers) == 1:
                resolved[col_key] = headers[0]
            else:
                resolved[col_key] = None  # ambiguous → blank
                warnings.append(
                    f"{len(headers)} columns mapped to '{col_key}'; left blank."
                )
        return resolved, warnings

    def _prepare(
        self, job: ImportJob, importer: ImporterDef
    ) -> Tuple[List[PreparedRow], List[dict], List[str]]:
        """Returns (valid prepared rows, errors[{row,column,message}], warnings).
        Pure — zero writes. Used by BOTH validate and commit (re-validation)."""
        max_rows, _ = self.caps(job.tenant_id)
        content = self._read_content(job)
        fmt = readers.sniff_format(content) or readers.FMT_CSV
        _, records = readers.read_rows(content, fmt, job.sheet_name, max_rows)
        col_to_header, warnings = self._invert_mapping(importer, job.mapping_json or {})
        assumed = self._assumed_tz(job)

        errors: List[dict] = []
        prepared: List[PreparedRow] = []
        id_header = col_to_header.get("id")  # the match key (D5)

        # First pass: coerce + per-cell validate; collect resolver inputs. Each
        # staged row carries its missing-required cols — enforced ONLY on create
        # rows in pass 2 (partial update for existing rows; required-create-only D5).
        staged: List[Tuple[int, dict, Optional[str], bool, list]] = []
        resolver_inputs: Dict[str, set] = {}
        seen_ids: set = set()
        unique_seen: Dict[str, set] = {c.key: set() for c in importer.columns if c.unique}

        for i, rec in enumerate(records):
            rownum = i + 2  # 1-based + header row
            data: dict = {}
            row_ok = True
            missing_required: list = []
            raw_id = (str(rec.get(id_header)).strip() if id_header and rec.get(id_header) is not None else "") or None

            for col in importer.columns:
                if col.key == "id":
                    continue
                header = col_to_header.get(col.key)
                raw = rec.get(header) if header else None
                value, err = coerce(raw, col)
                if err:
                    errors.append({"row": rownum, "column": col.key, "message": err})
                    row_ok = False
                    continue
                # required (if mapped-absent or empty) — deferred: only create rows
                if col.required and (value is None or value == "") and not col.resolver:
                    missing_required.append(col.key)
                    continue
                # explicit options membership (enum already coerced; non-enum w/ options)
                if value is not None and col.options and col.type != "enum":
                    allowed = {o["value"] if isinstance(o, dict) else o for o in (col.options if isinstance(col.options, list) else [])}
                    if allowed and value not in allowed:
                        errors.append({"row": rownum, "column": col.key, "message": "not an allowed value"})
                        row_ok = False
                        continue
                # validators
                for v in col.validators:
                    msg = v(value)
                    if msg:
                        errors.append({"row": rownum, "column": col.key, "message": msg})
                        row_ok = False
                        break
                else:
                    # datetime → UTC (D6)
                    if col.type == "datetime" and isinstance(value, datetime):
                        value = to_utc(value, assumed)
                    # unique in-file dup
                    if col.unique and value is not None:
                        if value in unique_seen[col.key]:
                            errors.append({"row": rownum, "column": col.key, "message": "duplicate value in file"})
                            row_ok = False
                        else:
                            unique_seen[col.key].add(value)
                    # resolver input collection
                    if col.resolver and value is not None:
                        if col.multi_value:
                            for item in str(value).split(","):
                                item = item.strip()
                                if item:
                                    resolver_inputs.setdefault(col.key, set()).add(item)
                        else:
                            resolver_inputs.setdefault(col.key, set()).add(value)
                    data[col.key] = value

            # imperative cross-column rule (D6)
            if importer.validate_row:
                for ckey, msg in (importer.validate_row(data, job.context_json or {}) or {}).items():
                    errors.append({"row": rownum, "column": ckey, "message": msg})
                    row_ok = False

            # in-file duplicate id
            if raw_id is not None:
                if raw_id in seen_ids:
                    errors.append({"row": rownum, "column": "id", "message": "duplicate id in file"})
                    row_ok = False
                seen_ids.add(raw_id)

            staged.append((rownum, data, raw_id, row_ok, missing_required))

        # Batched resolvers (set-based, ≤ one query per resolver column).
        resolved_maps: Dict[str, Dict[str, str]] = {}
        for col in importer.columns:
            if col.resolver and col.key in resolver_inputs:
                vals = list(resolver_inputs[col.key])
                resolved_maps[col.key] = col.resolver.lookup(self.db, job.tenant_id, vals) or {}

        # Batched match-existence (D7) for update/upsert id matching.
        existing_id_set: set = set()
        if importer.existing_ids and (job.mode in (MODE_UPDATE_ONLY, MODE_UPSERT)):
            ids = [sid for (_, _, sid, _, _) in staged if sid]
            if ids:
                existing_id_set = importer.existing_ids(self.db, job.tenant_id, ids)

        # Batched table-uniqueness (D6) — a `unique` column value already present
        # in the table must fail at VALIDATE (Test), not blow up the commit on a
        # DB UNIQUE constraint. Case-insensitive (matches the lower() transform +
        # the tenant-scoped uniqueness intent). One query per unique column.
        from sqlalchemy import func

        unique_existing: Dict[str, set] = {}
        for col in importer.columns:
            column = getattr(importer.model, col.attr, None)
            if not col.unique or column is None:
                continue
            values = {
                str(d[col.key]).lower()
                for (_, d, _, _, _) in staged
                if d.get(col.key) is not None
            }
            if not values:
                continue
            rows = (
                self.db.query(func.lower(column))
                .filter(
                    importer.model.tenant_id == job.tenant_id,
                    func.lower(column).in_(values),
                )
                .all()
            )
            unique_existing[col.key] = {r[0] for r in rows}

        # Second pass: apply resolvers + id/mode rules → prepared rows.
        for rownum, data, raw_id, row_ok, missing_required in staged:
            if not row_ok:
                continue
            # id / mode rules first (so required is enforced on create rows only).
            op, op_err = self._resolve_op(job.mode, raw_id, existing_id_set)
            if op_err:
                errors.append({"row": rownum, "column": "id", "message": op_err})
                continue
            if op == "create" and missing_required:
                for ckey in missing_required:
                    errors.append({"row": rownum, "column": ckey, "message": "required"})
                continue
            # table-uniqueness: a create row whose unique value already exists.
            if op == "create":
                dup_col = next(
                    (
                        col.key
                        for col in importer.columns
                        if col.unique
                        and data.get(col.key) is not None
                        and str(data[col.key]).lower() in unique_existing.get(col.key, set())
                    ),
                    None,
                )
                if dup_col:
                    errors.append(
                        {"row": rownum, "column": dup_col, "message": "a record with this value already exists"}
                    )
                    continue
            # resolver application
            resolver_failed = False
            for col in importer.columns:
                if not col.resolver or col.key not in data:
                    continue
                rmap = resolved_maps.get(col.key, {})
                value = data[col.key]
                if col.multi_value:
                    ids = []
                    for item in str(value).split(","):
                        item = item.strip()
                        if not item:
                            continue
                        rid = rmap.get(item)
                        if rid is None and col.resolver.mode != RESOLVE_FIND_OR_CREATE:
                            errors.append({"row": rownum, "column": col.key, "message": f"no match for '{item}'"})
                            resolver_failed = True
                        else:
                            ids.append(rid or item)  # find_or_create resolved at commit
                    data[col.key] = ids
                else:
                    rid = rmap.get(value)
                    if rid is None and col.resolver.mode != RESOLVE_FIND_OR_CREATE:
                        errors.append({"row": rownum, "column": col.key, "message": f"no match for '{value}'"})
                        resolver_failed = True
                    else:
                        data[col.key] = rid if rid is not None else value
            if resolver_failed:
                continue

            prepared.append(PreparedRow(rownum, data, raw_id, op))

        # Aggregate/set-based validation (sprint-4/05) — runs over the VALID set
        # only, zero writes, in both Test and commit. Rows it flags are demoted
        # out of `prepared` so the commit never acts on them (e.g. capacity).
        if importer.validate_prepared and prepared:
            agg_errors = (
                importer.validate_prepared(
                    self.db,
                    job.tenant_id,
                    [{"row": p.index, **p.data} for p in prepared],
                    job.context_json or {},
                )
                or []
            )
            if agg_errors:
                errors.extend(agg_errors)
                bad_rows = {e["row"] for e in agg_errors if e.get("row") is not None}
                if any(e.get("row") is None for e in agg_errors):
                    prepared = []  # an aggregate failure invalidates the whole set
                elif bad_rows:
                    prepared = [p for p in prepared if p.index not in bad_rows]

        return prepared, errors, warnings

    def _resolve_op(self, mode: str, raw_id: Optional[str], existing: set):
        if mode == MODE_CREATE_ONLY:
            if raw_id is not None:
                return None, "id must be blank for create-only"
            return "create", None
        if mode == MODE_UPDATE_ONLY:
            if raw_id is None:
                return None, "id is required for update-only"
            if raw_id not in existing:
                return None, "no record with this id"
            return "update", None
        # upsert
        if raw_id is None:
            return "create", None
        if raw_id in existing:
            return "update", None
        return None, "id provided but no record found"

    def _assumed_tz(self, job: ImportJob):
        if job.assumed_tz:
            try:
                return ZoneInfo(job.assumed_tz)
            except Exception:  # noqa: BLE001
                return timezone.utc
        return timezone.utc

    # ── validate (dry-run) ───────────────────────────────────────────────────

    def run_validate(self, db: Session, job_id: str) -> None:
        job = ImportRepository(db).get_unscoped(job_id)
        if job is None:
            return
        importer = get_importer(job.entity_type)
        try:
            prepared, errors, _warnings = self._prepare(job, importer)
            invalid_rownums = {e["row"] for e in errors}
            total = len(prepared) + len(invalid_rownums)
            job.total_rows = total
            job.valid_rows = len(prepared)
            job.invalid_rows = len(invalid_rownums)
            job.errors_json = errors[:ERRORS_CAP] or None
            if errors:
                job.error_report_key = self._build_error_file(job, importer, errors)
            job.status = STATUS_VALIDATED
            job.validated_at = datetime.now(timezone.utc)
            db.commit()
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            job = ImportRepository(db).get_unscoped(job_id)
            if job:
                job.status = STATUS_FAILED
                job.errors_json = [{"row": 0, "column": "", "message": f"Validation failed: {exc}"}]
                db.commit()

    # ── commit (all-or-nothing) ──────────────────────────────────────────────

    def run_commit(self, db: Session, job_id: str) -> None:
        repo = ImportRepository(db)
        if not repo.claim_for_commit(job_id, STATUS_VALIDATED):
            return  # already claimed (double-commit guard)
        db.commit()
        job = repo.get_unscoped(job_id)
        importer = get_importer(job.entity_type)
        try:
            prepared, errors, _warnings = self._prepare(job, importer)
            if job.abort_on_invalid and errors:
                raise RuntimeError(f"{len(errors)} invalid rows — commit aborted")

            creates = [p for p in prepared if p.op == "create"]
            updates = [p for p in prepared if p.op == "update"]
            created_ids: List[str] = []
            updated_ids: List[str] = []
            # Capture before-images for update rows so we can emit a field-level
            # `changes` diff (D13) — `entity.field_changed`/`updated` workflows
            # need it. Done BEFORE update_rows mutates.
            update_changes: Dict[str, Dict[str, dict]] = {}
            old_vals: Dict[str, dict] = {}
            if updates:
                upd_ids = [p.record_id for p in updates]
                before = {
                    r.id: r
                    for r in db.query(importer.model)
                    .filter(importer.model.id.in_(upd_ids))
                    .all()
                }
                for p in updates:
                    rec = before.get(p.record_id)
                    if rec is None:
                        continue
                    old_vals[p.record_id] = {
                        col.attr: getattr(rec, col.attr, None)
                        for col in importer.columns
                        if col.key != "id" and col.key in p.data
                    }
            if creates and importer.create_rows:
                created_ids = importer.create_rows(
                    db, job.tenant_id, [p.data for p in creates], job.context_json or {}
                )
            if updates and importer.update_rows:
                updated_ids = importer.update_rows(
                    db,
                    job.tenant_id,
                    [{**p.data, "id": p.record_id} for p in updates],
                    job.context_json or {},
                )
                for p in updates:
                    ov = old_vals.get(p.record_id, {})
                    ch = {}
                    for col in importer.columns:
                        if col.key == "id" or col.key not in p.data:
                            continue
                        new = p.data[col.key]
                        old = ov.get(col.attr)
                        if str(old) != str(new):
                            ch[col.attr] = {"from": old, "to": new}
                    if ch:
                        update_changes[p.record_id] = ch
            job.created_ids = created_ids or None
            job.updated_ids = updated_ids or None
            job.valid_rows = len(prepared)
            job.invalid_rows = len({e["row"] for e in errors})
            job.errors_json = errors[:ERRORS_CAP] or None
            job.status = STATUS_DONE
            job.committed_at = datetime.now(timezone.utc)
            job.finished_at = datetime.now(timezone.utc)
            db.commit()  # ONE transaction over the whole valid set

            if job.trigger_automations:
                self._emit_events(db, job, importer, created_ids, updated_ids, update_changes)
        except Exception as exc:  # noqa: BLE001 — D7: roll back EVERYTHING
            db.rollback()
            job = repo.get_unscoped(job_id)
            if job:
                job.status = STATUS_FAILED
                job.finished_at = datetime.now(timezone.utc)
                existing = job.errors_json or []
                job.errors_json = ([{"row": 0, "column": "", "message": f"Commit failed: {exc}"}] + existing)[:ERRORS_CAP]
                db.commit()

    def _emit_events(self, db, job, importer, created_ids, updated_ids, update_changes=None):
        """D13 ON-path: per-row entity.created / entity.updated AFTER the commit.
        The import commits internally, so this uses the emit + explicit
        dispatch_pending pattern (not the after-commit drain). Failure-isolated:
        a broken/slow workflow never breaks the import (dispatch_pending swallows).
        ``actor_id`` (a string id), not ``actor`` (which expects a User)."""
        try:
            from app.workflow_engine.entities import get_workflow_entity, load_record
            from app.workflow_engine.entity_events import (
                dispatch_pending,
                emit_entity_event,
            )

            wf_entity = get_workflow_entity(job.entity_type)
            if wf_entity is None:
                return
            changes_map = update_changes or {}
            for rid in created_ids or []:
                rec = load_record(db, wf_entity, job.tenant_id, rid)
                if rec is not None:
                    emit_entity_event(
                        db, job.entity_type, "created", rec,
                        tenant_id=job.tenant_id, actor_id=job.actor_user_id,
                    )
            for rid in updated_ids or []:
                rec = load_record(db, wf_entity, job.tenant_id, rid)
                if rec is not None:
                    # field-level changes → entity.field_changed + updated workflows.
                    emit_entity_event(
                        db, job.entity_type, "updated", rec,
                        tenant_id=job.tenant_id, actor_id=job.actor_user_id,
                        changes=changes_map.get(rid),
                    )
            dispatch_pending(db)  # drain the buffer NOW → match + run workflows
        except Exception:  # noqa: BLE001
            logger.exception("import workflow dispatch failed for job %s", job.id)

    # ── template + error-file builders (sanitized cells, D14) ─────────────────

    def build_template(self, entity_type: str, columns: List[str], fmt: str) -> Tuple[bytes, str, str]:
        importer = self._require_importer(entity_type)
        chosen = [c for c in importer.columns if c.required or c.key in columns]
        # id first if present (export↔import symmetry, D12) — but template for
        # create doesn't need it; include only if explicitly chosen.
        headers = [c.label for c in chosen]
        if fmt == "csv":
            import csv as _csv

            buf = io.StringIO()
            w = _csv.writer(buf)
            w.writerow([sanitize_cell(h) for h in headers])
            return buf.getvalue().encode("utf-8"), "text/csv", "csv"
        return self._build_xlsx_template(chosen, headers)

    def _build_xlsx_template(self, chosen: List[ImportColumn], headers: List[str]):
        import openpyxl
        from openpyxl.styles import Font
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import"
        for col_idx, (col, header) in enumerate(zip(chosen, headers), start=1):
            cell = ws.cell(row=1, column=col_idx, value=sanitize_cell(header))
            cell.font = Font(bold=True)
            # in-cell dropdown for bounded enum/option columns
            opts = col.options if isinstance(col.options, list) else None
            if opts:
                values = ",".join(str(o["value"] if isinstance(o, dict) else o) for o in opts)
                if len(values) <= 255:
                    dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=not col.required)
                    ws.add_data_validation(dv)
                    letter = openpyxl.utils.get_column_letter(col_idx)
                    dv.add(f"{letter}2:{letter}1000")
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"

    def _build_error_file(self, job: ImportJob, importer: ImporterDef, errors: List[dict]) -> str:
        """Annotated error file = the COMPLETE error set (original row + _error)."""
        import csv as _csv

        by_row: Dict[int, List[str]] = {}
        for e in errors:
            by_row.setdefault(e["row"], []).append(f"{e['column']}: {e['message']}")
        buf = io.StringIO()
        w = _csv.writer(buf)
        w.writerow(["row", "_error"])
        for rownum in sorted(by_row):
            w.writerow([rownum, sanitize_cell("; ".join(by_row[rownum]))])
        store = storage_for_tenant(self.db, job.tenant_id)
        return store.save(
            f"imports/{job.id}/errors.csv", buf.getvalue().encode("utf-8"), "text/csv"
        )

    def error_file(self, tenant_id: str, job_id: str) -> Tuple[bytes, str]:
        job = self._require_job(tenant_id, job_id)
        if not job.error_report_key:
            raise HTTPException(404, "No error file.")
        store = storage_for_tenant(self.db, tenant_id)
        kind, value = store.resolve(job.error_report_key)
        if kind == "path":
            with open(value, "rb") as fh:
                return fh.read(), "errors.csv"
        import urllib.request

        return urllib.request.urlopen(value).read(), "errors.csv"  # noqa: S310

    # ── mapping preview (sheets + headers + auto-map) ────────────────────────

    @staticmethod
    def _normalize(s: str) -> str:
        return "".join(ch for ch in str(s).lower() if ch.isalnum())

    def preview(self, tenant_id: str, job_id: str, sheet: Optional[str] = None) -> dict:
        """Detected sheets + the chosen sheet's headers + an auto-mapping
        (normalized header → catalog column). A clean template round-trip
        pre-selects 100% (D4); the user overrides freely on the mapping page."""
        job = self._require_job(tenant_id, job_id)
        importer = self._require_importer(job.entity_type)
        content = self._read_content(job)
        fmt = readers.sniff_format(content) or readers.FMT_CSV
        sheets = readers.list_sheets(content, fmt)
        chosen = sheet or job.sheet_name or (sheets[0] if sheets else None)
        max_rows, _ = self.caps(tenant_id)
        headers, _records = readers.read_rows(content, fmt, chosen, max_rows)
        by_norm = {self._normalize(c.label): c.key for c in importer.columns}
        by_norm.update({self._normalize(c.key): c.key for c in importer.columns})
        auto = {h: by_norm.get(self._normalize(h)) for h in headers}
        return {"sheets": sheets, "sheetName": chosen, "headers": headers, "autoMapping": auto}

    # ── reads ────────────────────────────────────────────────────────────────

    def get_job(self, tenant_id: str, job_id: str) -> ImportJob:
        return self._require_job(tenant_id, job_id)

    def list_jobs(self, tenant_id, **kw):
        return self.repo.list(tenant_id, **kw)

    # ── helpers ──────────────────────────────────────────────────────────────

    def _require_importer(self, entity_type: str) -> ImporterDef:
        importer = get_importer(entity_type)
        if importer is None:
            raise HTTPException(404, f"No importer for '{entity_type}'.")
        return importer

    def _require_job(self, tenant_id: str, job_id: str) -> ImportJob:
        job = self.repo.get(tenant_id, job_id)
        if job is None:
            raise HTTPException(404, "Import job not found.")
        return job
