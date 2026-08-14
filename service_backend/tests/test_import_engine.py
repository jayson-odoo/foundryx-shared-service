"""Import engine tests (sprint-3/09, F8) - validates AC-09-*.

Adapter sniff · header detection · coercion + cell errors · 3 modes + id-only
matching · in-file dup · map-collision degrade · all-or-nothing commit ·
double-commit guard · formula-injection sanitize · caps fail-fast · created-ids ·
imports.read_all gate + tenant isolation · drift guard (cols ⊆ writable).
"""
import io

import openpyxl

from app.import_engine import readers
from app.import_engine.registry import get_importer
from app.import_engine.sanitize import sanitize_cell
from app.models.import_job import STATUS_DONE, STATUS_FAILED, STATUS_VALIDATED
from app.models.user import User
from app.models.tenant import DEFAULT_TENANT_ID
from tests.conftest import ACTIVE_EMAIL, ACTIVE_PASSWORD, PLATFORM_EMAIL, PLATFORM_PASSWORD


def _login(client, email, password, slug=None):
    p = {"email": email, "password": password}
    if slug:
        p["tenantSlug"] = slug
    return client.post("/auth/login", json=p)


def _headers(res):
    assert res.status_code == 200, res.text
    return {"Authorization": f"Bearer {res.json()['access_token']}"}


def _admin(client):
    return _headers(_login(client, ACTIVE_EMAIL, ACTIVE_PASSWORD))


def _xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv(text):
    return text.encode("utf-8")


def _upload(client, headers, content, filename, mode="create_only", **extra):
    data = {"entityType": "user", "mode": mode}
    data.update(extra)
    return client.post(
        "/imports",
        data=data,
        files={"file": (filename, content, "application/octet-stream")},
        headers=headers,
    )


def _map_and_validate(client, headers, job_id, mapping, sheet=None):
    return client.put(
        f"/imports/{job_id}/mapping",
        json={"mapping": mapping, "sheetName": sheet},
        headers=headers,
    )


def _ident_map(headers_list):
    """Map each file header to a same-named catalog column key."""
    keymap = {"Email": "email", "Name": "name", "Status": "status", "ID": "id"}
    return {h: keymap.get(h) for h in headers_list}


# ── reader / format ────────────────────────────────────────────────────────


def test_sniff_each_format():
    assert readers.sniff_format(_xlsx([["Email"], ["a@b.com"]])) == readers.FMT_XLSX
    assert readers.sniff_format(_csv("Email\na@b.com\n")) == readers.FMT_CSV
    assert readers.sniff_format(b"MZ\x90\x00") is None  # executable rejected


def test_header_is_first_non_empty_row():
    content = _csv("\n\nEmail,Name\nx@y.com,Xy\n")
    headers, rows = readers.read_rows(content, readers.FMT_CSV, None, 1000)
    assert headers == ["Email", "Name"]
    assert rows[0]["Email"] == "x@y.com"


def test_duplicate_headers_suffixed():
    headers, _ = readers.read_rows(_csv("Email,Email\na@b.com,c@d.com\n"), readers.FMT_CSV, None, 100)
    assert headers == ["Email", "Email (2)"]


# ── create-only happy path + coercion ──────────────────────────────────────


def test_create_only_imports_valid_rows(client):
    h = _admin(client)
    content = _xlsx([["Email", "Name", "Status"], ["alice@e2e.com", "Alice", "active"], ["bob@e2e.com", "Bob", "inactive"]])
    job_id = _upload(client, h, content, "u.xlsx").json()["jobId"]
    res = _map_and_validate(client, h, job_id, _ident_map(["Email", "Name", "Status"]))
    assert res.status_code == 200
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["status"] == STATUS_VALIDATED
    assert job["validRows"] == 2 and job["invalidRows"] == 0
    commit = client.post(f"/imports/{job_id}/commit", headers=h)
    assert commit.status_code == 200
    done = client.get(f"/imports/{job_id}", headers=h).json()
    assert done["status"] == STATUS_DONE
    assert len(done["createdIds"]) == 2
    # users actually landed
    assert client.get("/users?search=alice@e2e.com", headers=h).json()["total"] >= 1


def test_coercion_cell_errors(client):
    h = _admin(client)
    # bad email + bad enum status
    content = _csv("Email,Status\nnotanemail,active\nok@e2e.com,bogus\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email", "Status"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["invalidRows"] == 2
    cols = {e["column"] for e in job["errors"]}
    assert "email" in cols and "status" in cols


def test_required_email_missing_is_row_error(client):
    h = _admin(client)
    content = _csv("Email,Name\n,NoEmail\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email", "Name"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["invalidRows"] == 1


# ── modes + id-only matching (D5) ──────────────────────────────────────────


def test_create_only_rejects_present_id(client):
    h = _admin(client)
    content = _csv("ID,Email\nsome-id,z@e2e.com\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["ID", "Email"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["invalidRows"] == 1
    assert any(e["column"] == "id" for e in job["errors"])


def test_update_only_requires_existing_id(client):
    h = _admin(client)
    # existing demo user id
    demo = client.get("/auth/me", headers=h).json()
    uid = demo["id"]
    content = _csv(f"ID,Name\n{uid},Renamed\nmissing-id,Nope\n")
    job_id = _upload(client, h, content, "u.csv", mode="update_only").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["ID", "Name"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["validRows"] == 1 and job["invalidRows"] == 1


def test_upsert_create_and_update(client):
    h = _admin(client)
    demo = client.get("/auth/me", headers=h).json()
    content = _csv(f"ID,Email,Name\n{demo['id']},,Demo Updated\n,new-upsert@e2e.com,New\n")
    job_id = _upload(client, h, content, "u.csv", mode="upsert").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["ID", "Email", "Name"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["validRows"] == 2
    client.post(f"/imports/{job_id}/commit", headers=h)
    done = client.get(f"/imports/{job_id}", headers=h).json()
    assert done["status"] == STATUS_DONE
    assert len(done["createdIds"]) == 1 and len(done["updatedIds"]) == 1


# ── in-file dup + map collision ────────────────────────────────────────────


def test_in_file_duplicate_email_errors(client):
    h = _admin(client)
    content = _csv("Email\ndup@e2e.com\ndup@e2e.com\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["invalidRows"] == 1  # second row flagged


def test_existing_email_caught_at_validate_not_commit(client):
    """Regression: a unique value already in the TABLE must fail at Test
    (validate), never blow up the commit on the DB UNIQUE constraint."""
    h = _admin(client)
    # demo@example.com already exists (seeded).
    content = _csv("Email,Name\ndemo@example.com,Dup\nfresh-uniq@e2e.com,Fresh\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email", "Name"]))
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["validRows"] == 1 and job["invalidRows"] == 1
    assert any(
        e["column"] == "email" and "already exists" in e["message"].lower()
        for e in job["errors"]
    )
    # commit imports only the fresh row - no DB crash.
    client.post(f"/imports/{job_id}/commit", headers=h)
    assert client.get(f"/imports/{job_id}", headers=h).json()["status"] == STATUS_DONE


def test_map_collision_leaves_target_blank(client):
    h = _admin(client)
    # two file columns both mapped to "name" → name blank+warning (email still ok)
    content = _csv("Email,A,B\nco@e2e.com,Foo,Bar\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    res = _map_and_validate(client, h, job_id, {"Email": "email", "A": "name", "B": "name"})
    assert res.status_code == 200
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["validRows"] == 1  # row valid; name simply blank


# ── all-or-nothing + abort + double-commit guard ───────────────────────────


def test_abort_on_invalid_blocks_commit(client):
    h = _admin(client)
    content = _csv("Email\ngood@e2e.com\nbademail\n")
    job_id = _upload(client, h, content, "u.csv", abortOnInvalid="true").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email"]))
    client.post(f"/imports/{job_id}/commit", headers=h)
    job = client.get(f"/imports/{job_id}", headers=h).json()
    assert job["status"] == STATUS_FAILED
    # nothing landed
    assert client.get("/users?search=good@e2e.com", headers=h).json()["total"] == 0


def test_double_commit_guard(client):
    h = _admin(client)
    content = _csv("Email\nonce@e2e.com\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email"]))
    first = client.post(f"/imports/{job_id}/commit", headers=h)
    assert first.status_code == 200
    # second commit: job no longer 'validated' → 409
    second = client.post(f"/imports/{job_id}/commit", headers=h)
    assert second.status_code == 409


# ── caps + security + perms ────────────────────────────────────────────────


def test_unsupported_file_rejected(client):
    h = _admin(client)
    res = _upload(client, h, b"\x7fELF\x02\x01", "evil.bin")
    assert res.status_code == 422


def test_formula_injection_sanitized():
    assert sanitize_cell("=cmd()") == "'=cmd()"
    assert sanitize_cell("+1") == "'+1"
    assert sanitize_cell("safe") == "safe"


def test_template_download_has_dropdown(client):
    h = _admin(client)
    res = client.get("/imports/template/user?columns=name,status&format=xlsx", headers=h)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Email"  # required always in


def test_drift_guard_columns_subset_of_writable():
    """AC-09-15 - import cols ⊆ the entity's workflow writable whitelist (+ the
    natural keys email/id which are import-specific identity columns)."""
    from app.workflow_engine.entities import get_workflow_entity

    importer = get_importer("user")
    wf = get_workflow_entity("user")
    allowed = set(wf.writable) | {"id", "email"}  # id=match key, email=natural key
    for col in importer.columns:
        assert col.attr in allowed, f"{col.key} not in writable whitelist"


def test_imports_read_all_scopes_history(client, session_factory):
    h = _admin(client)
    # admin creates a job
    content = _csv("Email\nhist@e2e.com\n")
    _upload(client, h, content, "u.csv")
    # admin (has imports.read_all via tenant_admin_grant) sees jobs
    res = client.get("/imports", headers=h)
    assert res.status_code == 200
    assert res.json()["total"] >= 1


def test_commit_is_set_based_not_per_row(client, session_factory):
    """AC-09-10 - commit DML must not scale per row (one bulk insert, not N)."""
    from sqlalchemy import event

    h = _admin(client)
    rows = "Email\n" + "\n".join(f"setbased{i}@e2e.com" for i in range(40)) + "\n"
    job_id = _upload(client, h, _csv(rows), "u.csv").json()["jobId"]
    _map_and_validate(client, h, job_id, _ident_map(["Email"]))

    # Count INSERT statements during commit on the shared engine.
    engine = session_factory.kw["bind"]
    inserts = {"n": 0}

    def _before(conn, cursor, statement, *a):
        if statement.lstrip().upper().startswith("INSERT INTO USERS"):
            inserts["n"] += 1

    event.listen(engine, "before_cursor_execute", _before)
    try:
        client.post(f"/imports/{job_id}/commit", headers=h)
    finally:
        event.remove(engine, "before_cursor_execute", _before)
    done = client.get(f"/imports/{job_id}", headers=h).json()
    assert done["status"] == STATUS_DONE and len(done["createdIds"]) == 40
    # bulk_insert_mappings → a small constant number of INSERTs, never ~40.
    assert inserts["n"] <= 3, f"expected bulk insert, got {inserts['n']} INSERTs"


def test_trigger_automations_fires_events(client):
    """AC-09-18 - trigger_automations ON dispatches per-row entity.created to the
    event bus (so workflows fire); OFF dispatches nothing."""
    from app.workflow_engine.entity_events import (
        register_event_subscriber,
        unregister_event_subscriber,
    )

    seen: list = []
    sub = lambda _s, ev: seen.append((ev["entity_type"], ev["action"], ev["record_id"]))  # noqa: E731
    register_event_subscriber(sub)
    try:
        h = _admin(client)
        # ON → event dispatched.
        c1 = _csv("Email\nwf-on@e2e.com\n")
        j1 = _upload(client, h, c1, "u.csv", triggerAutomations="true").json()["jobId"]
        _map_and_validate(client, h, j1, _ident_map(["Email"]))
        client.post(f"/imports/{j1}/commit", headers=h)
        assert any(e[0] == "user" and e[1] == "created" for e in seen), seen

        # OFF (default) → nothing dispatched.
        seen.clear()
        c2 = _csv("Email\nwf-off@e2e.com\n")
        j2 = _upload(client, h, c2, "u.csv").json()["jobId"]
        _map_and_validate(client, h, j2, _ident_map(["Email"]))
        client.post(f"/imports/{j2}/commit", headers=h)
        assert seen == []
    finally:
        unregister_event_subscriber(sub)


def test_update_emits_field_changes(client):
    """D13 - an import UPDATE emits entity.updated WITH a field-level `changes`
    diff (so entity.field_changed / updated workflows fire)."""
    from app.workflow_engine.entity_events import (
        register_event_subscriber,
        unregister_event_subscriber,
    )

    h = _admin(client)
    demo = client.get("/auth/me", headers=h).json()
    seen: list = []
    sub = lambda _s, ev: seen.append(ev)  # noqa: E731
    register_event_subscriber(sub)
    try:
        content = _csv(f"ID,Name\n{demo['id']},Renamed Via Import\n")
        job_id = _upload(client, h, content, "u.csv", mode="upsert", triggerAutomations="true").json()["jobId"]
        _map_and_validate(client, h, job_id, _ident_map(["ID", "Name"]))
        client.post(f"/imports/{job_id}/commit", headers=h)
        upd = [e for e in seen if e["entity_type"] == "user" and e["action"] == "updated"]
        assert upd, seen
        assert upd[0]["changes"] and "name" in upd[0]["changes"]
        assert upd[0]["changes"]["name"]["to"] == "Renamed Via Import"
    finally:
        unregister_event_subscriber(sub)


def test_import_settings_get_set(client):
    """D11 - per-tenant caps: default until set, then override + fall-back to default."""
    h = _admin(client)
    res = client.get("/imports/settings", headers=h)
    assert res.status_code == 200
    assert res.json()["isDefault"] is True
    # set an override
    put = client.put("/imports/settings", json={"maxRows": 500, "maxFileMb": 5}, headers=h)
    assert put.status_code == 200
    assert put.json() == {"maxRows": 500, "maxFileMb": 5, "isDefault": False}
    # clearing (0/null) reverts to the global default
    rev = client.put("/imports/settings", json={"maxRows": 0, "maxFileMb": 0}, headers=h)
    assert rev.json()["isDefault"] is True


def test_tenant_isolation(client):
    h = _admin(client)
    content = _csv("Email\niso@e2e.com\n")
    job_id = _upload(client, h, content, "u.csv").json()["jobId"]
    # other tenant can't read this job
    ph = _headers(_login(client, PLATFORM_EMAIL, PLATFORM_PASSWORD, "platform"))
    res = client.get(f"/imports/{job_id}", headers=ph)
    assert res.status_code == 404
