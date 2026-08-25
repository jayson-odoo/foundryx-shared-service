"""Multi-format reader adapter (sprint-3/09 D10/D14) - one magic-byte-sniffing
entry, format-blind downstream.

Accepts xlsx/xlsm (openpyxl, ``read_only=True, data_only=True`` - streams values
not formulas, stops at the cap), xls (xlrd, legacy whole-file path bounded by the
size cap), csv (stdlib + charset-normalizer). ``.xlsm`` macros never execute
(cell-read only). Extension/declared content-type are hints; magic bytes gate.

Header detection is POSITIONAL (D4): skip leading fully-empty rows, the first row
with content = headers, data follows. Duplicate headers auto-suffix ("Email",
"Email (2)") so each maps independently.
"""
import csv as _csv
import io
from typing import Dict, List, Optional, Tuple

FMT_XLSX = "xlsx"
FMT_XLS = "xls"
FMT_CSV = "csv"


def sniff_format(content: bytes) -> Optional[str]:
    """Magic-byte format gate. xlsx/xlsm = zip (PK); xls = OLE2 compound; else
    try csv (decodable text). Returns the format or None (unsupported)."""
    if content[:4] in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08"):
        return FMT_XLSX
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return FMT_XLS
    # Block executables / markup masquerading as csv.
    if content[:2] == b"MZ" or content[:4] == b"\x7fELF":
        return None
    head = content[:512].lstrip().lower()
    if head.startswith((b"<svg", b"<?xml", b"<!doctype", b"<html")):
        return None
    if _decode(content) is not None:
        return FMT_CSV
    return None


def _decode(content: bytes) -> Optional[str]:
    try:
        from charset_normalizer import from_bytes

        match = from_bytes(content).best()
        if match is not None:
            return str(match)
    except Exception:  # noqa: BLE001
        pass
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        return None


def list_sheets(content: bytes, fmt: str) -> List[str]:
    if fmt == FMT_XLSX:
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        names = wb.sheetnames
        wb.close()
        return names
    if fmt == FMT_XLS:
        import xlrd

        book = xlrd.open_workbook(file_contents=content)
        return book.sheet_names()
    return ["Sheet1"]


def _suffix_headers(raw: List[str]) -> List[str]:
    """Auto-suffix duplicate headers: Email, Email (2), Email (3) (D4)."""
    seen: Dict[str, int] = {}
    out: List[str] = []
    for h in raw:
        name = (str(h).strip() if h is not None else "") or "Column"
        seen[name] = seen.get(name, 0) + 1
        out.append(name if seen[name] == 1 else f"{name} ({seen[name]})")
    return out


def _rows_to_records(
    rows: List[List], max_rows: int
) -> Tuple[List[str], List[dict]]:
    """Positional header detection + record build. ``rows`` = raw cell matrix."""
    # Skip leading fully-empty rows; first non-empty = headers.
    start = 0
    while start < len(rows) and _is_empty_row(rows[start]):
        start += 1
    if start >= len(rows):
        return [], []
    headers = _suffix_headers(rows[start])
    records: List[dict] = []
    for raw in rows[start + 1 :]:
        if _is_empty_row(raw):
            continue
        rec = {}
        for i, header in enumerate(headers):
            rec[header] = raw[i] if i < len(raw) else None
        records.append(rec)
        if len(records) >= max_rows:
            break
    return headers, records


def _is_empty_row(row: List) -> bool:
    return all(
        c is None or (isinstance(c, str) and c.strip() == "") for c in (row or [])
    )


def read_rows(
    content: bytes, fmt: str, sheet_name: Optional[str], max_rows: int
) -> Tuple[List[str], List[dict]]:
    """Uniform (headers, list[dict]) for any format. Stops at ``max_rows`` data
    rows (the cap bounds memory; the size cap bounds the xls whole-file path)."""
    if fmt == FMT_XLSX:
        return _read_xlsx(content, sheet_name, max_rows)
    if fmt == FMT_XLS:
        return _read_xls(content, sheet_name, max_rows)
    return _read_csv(content, max_rows)


def _read_xlsx(content, sheet_name, max_rows):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows: List[List] = []
    # Read a bounded window: header search slack + max_rows data + 1.
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(row))
        if len(rows) >= max_rows + 50:
            break
    wb.close()
    return _rows_to_records(rows, max_rows)


def _read_xls(content, sheet_name, max_rows):
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = (
        book.sheet_by_name(sheet_name)
        if sheet_name and sheet_name in book.sheet_names()
        else book.sheet_by_index(0)
    )
    rows = [
        [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        for r in range(min(sheet.nrows, max_rows + 50))
    ]
    return _rows_to_records(rows, max_rows)


def _read_csv(content, max_rows):
    text = _decode(content) or ""
    reader = _csv.reader(io.StringIO(text))
    rows = []
    for i, row in enumerate(reader):
        rows.append(row)
        if len(rows) >= max_rows + 50:
            break
    return _rows_to_records(rows, max_rows)
